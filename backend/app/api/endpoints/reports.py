import os
import uuid
import json
import logging
from datetime import datetime, date
from io import BytesIO

import aiofiles
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Query, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc
from sqlalchemy.orm import joinedload
from pydantic import ValidationError

from app.core.database import get_db
from app.api.dependencies import get_current_user
from app.utils.bulk_import import process_file

from app.models.user import User
from app.models.organization import Organization
from app.models.report_submission import ReportSubmission
from app.models.investment_fact import InvestmentFact
from app.models.investment_forecast import InvestmentForecast
from app.models.uploaded_file import UploadedFile
from app.models.audit_log import AuditLog

router = APIRouter()
logger = logging.getLogger(__name__)

UPLOAD_PATH = os.environ.get('UPLOAD_PATH', './uploads')
os.makedirs(UPLOAD_PATH, exist_ok=True)

QUARTER_MONTHS = {1: "январь-март", 2: "апрель-июнь", 3: "июль-сентябрь", 4: "октябрь-декабрь"}

# Стили Excel
HEADER_FILL = PatternFill(start_color='5C6BC0', end_color='5C6BC0', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
MONEY_FORMAT = '#,##0.00" тыс. ₽"'
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def apply_money(cell):
    cell.number_format = MONEY_FORMAT
    cell.border = THIN_BORDER

# Mock данные
YEARLY_DATA = [{"year": 2022, "fact": 390509, "plan": 393401}, {"year": 2023, "fact": 420000, "plan": 410000}, {"year": 2024, "fact": 450000, "plan": 440000}, {"year": 2025, "fact": 384379, "plan": 470000}]

DISTRICTS_DATA = [
    {"name": "г. Тюмень", "fact": 169154, "plan": 170000, "orgs": 89}, {"name": "Тюменский район", "fact": 51465, "plan": 52000, "orgs": 28},
    {"name": "г. Ишим", "fact": 22902, "plan": 23000, "orgs": 12}, {"name": "г. Тобольск", "fact": 19551, "plan": 20000, "orgs": 18}
]

@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...), 
    report_type: str = Form(default="annual"), 
    year: int = Form(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)  # 🔒 Обязательно берем текущего юзера
):
    if year is None: 
        year = date.today().year
    
    quarter = None
    if report_type.startswith("q") and report_type[1].isdigit():
        quarter = int(report_type[1])

    # 1. Генерируем безопасное UUID имя файла и сохраняем его физически
    file_ext = file.filename.split('.')[-1] if '.' in file.filename else 'xlsx'
    stored_filename = f"{uuid.uuid4()}.{file_ext}"
    path = os.path.join(UPLOAD_PATH, stored_filename)
    
    content = await file.read()
    file_size = len(content)
    
    async with aiofiles.open(path, 'wb') as f:
        await f.write(content)
        
    # 2. Создаем запись о файле со статусом 'processing'
    uploaded_file_record = UploadedFile(
        organization_id=current_user.organization_id,
        original_filename=file.filename,
        stored_filename=stored_filename,
        file_type=file_ext,
        file_size_bytes=file_size,
        year=year,
        quarter=quarter,
        processing_status="processing",
        uploaded_by_user_id=current_user.id
    )
    db.add(uploaded_file_record)
    await db.flush() # flush, чтобы получить ID файла для AuditLog

    # Базовые данные для лога
    audit_details = {
        "original_filename": file.filename,
        "stored_filename": stored_filename,
        "file_size": file_size,
        "report_type": report_type,
        "year": year,
        "quarter": quarter
    }

    try:
        # 3. Парсинг и валидация через твой bulk_import.py
        import_results = await process_file(path, year, quarter, db)
        
        # Если process_file возвращает массив errors (наша бизнес-логика валидации)
        if import_results.get("errors"):
            uploaded_file_record.processing_status = "error"
            # Сохраняем ошибки как JSON строку
            uploaded_file_record.error_message = json.dumps(import_results["errors"], ensure_ascii=False)
            
            audit_details["status"] = "error_validation"
            audit_details["errors_count"] = len(import_results["errors"])
        else:
            # 4. Все отлично, данные валидны
            uploaded_file_record.processing_status = "success"
            audit_details["status"] = "success"
            audit_details["records_inserted"] = import_results.get("success", 0)
            
            # Обновляем статус в таблице ReportSubmissions (отчет сдан)
            if current_user.organization_id:
                stmt = select(ReportSubmission).where(
                    ReportSubmission.organization_id == current_user.organization_id,
                    ReportSubmission.year == year,
                    ReportSubmission.quarter == (quarter or 4) # Если annual, обычно привязан к 4 кварталу
                )
                submission = (await db.execute(stmt)).scalars().first()
                if submission:
                    submission.status = "submitted"
                    submission.submitted_date = date.today()

        # 5. Записываем действие в Аудит лог
        audit_log = AuditLog(
            user_id=current_user.id,
            action="upload_file",
            entity_type="uploaded_file",
            entity_id=uploaded_file_record.id,
            details=audit_details
        )
        db.add(audit_log)
        await db.commit()

        return {
            "status": uploaded_file_record.processing_status,
            "file_id": uploaded_file_record.id,
            "records_inserted": import_results.get("success", 0),
            "errors": import_results.get("errors", [])
        }

    except ValidationError as ve:
        # Перехват жестких ошибок Pydantic (если process_file прокидывает их наверх)
        await db.rollback() # Откатываем транзакцию фактов инвестиций
        
        errors = ve.errors()
        uploaded_file_record.processing_status = "error"
        uploaded_file_record.error_message = json.dumps(errors, ensure_ascii=False)
        
        audit_details["status"] = "error_pydantic"
        audit_details["exception"] = str(errors)
        
        # Пишем в лог, что загрузка провалилась
        audit_log = AuditLog(user_id=current_user.id, action="upload_file_failed", entity_type="uploaded_file", entity_id=uploaded_file_record.id, details=audit_details)
        db.add(uploaded_file_record)
        db.add(audit_log)
        await db.commit()
        
        return {"status": "error", "file_id": uploaded_file_record.id, "errors": errors}

    except Exception as e:
        # Любые другие краши (БД упала, нет колонок в Excel и тд)
        await db.rollback()
        logger.error(f"Critical error processing file: {str(e)}")
        
        uploaded_file_record.processing_status = "error"
        uploaded_file_record.error_message = json.dumps({"system_error": str(e)}, ensure_ascii=False)
        
        audit_details["status"] = "system_error"
        audit_details["exception"] = str(e)
        
        audit_log = AuditLog(user_id=current_user.id, action="upload_file_crashed", entity_type="uploaded_file", entity_id=uploaded_file_record.id, details=audit_details)
        db.add(uploaded_file_record)
        db.add(audit_log)
        await db.commit()
        
        raise HTTPException(status_code=500, detail="Системная ошибка при обработке файла. Информация передана администратору.")
@router.get("/template/{report_type}")
async def download_template(report_type: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    headers = ['Наименование', 'ИНН', 'ОКВЭД', 'ОКПО', 'СМП', 'Инвестиции, тыс. ₽']
    for col, h in enumerate(headers, 1): ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))
    ws.cell(row=2, column=1, value='ООО "Пример"')
    ws.cell(row=2, column=6, value=1000)
    apply_money(ws.cell(row=2, column=6))
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=template_{report_type}.xlsx"})

@router.get("/export/yearly")
async def export_yearly():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "По годам"
    ws.merge_cells('A1:D1')
    ws['A1'] = "Динамика инвестиций 2022-2025"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}"
    headers = ['Год', 'ФАКТ, тыс. ₽', 'ПЛАН, тыс. ₽', 'Освоение, %']
    for col, h in enumerate(headers, 1): ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 4)
    for i, d in enumerate(YEARLY_DATA, 5):
        ws.cell(row=i, column=1, value=d["year"])
        apply_money(ws.cell(row=i, column=2, value=d["fact"]))
        apply_money(ws.cell(row=i, column=3, value=d["plan"]))
        ws.cell(row=i, column=4, value=f"{round(d['fact']/d['plan']*100,1) if d['plan'] else 0}%")
    for col in [1, 2, 3, 4]: ws.column_dimensions[get_column_letter(col)].width = 20
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=yearly_2022-2025.xlsx"})

@router.get("/export/districts")
async def export_districts(year: int = Query(default=2022)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Районы {year}"
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Инвестиции по районам Тюменской области за {year} год"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y')}"
    headers = ['Район', 'Организаций', 'ФАКТ, тыс. ₽', 'ПЛАН, тыс. ₽', 'Освоение, %']
    for col, h in enumerate(headers, 1): ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 5)
    for i, d in enumerate(DISTRICTS_DATA, 5):
        ws.cell(row=i, column=1, value=d["name"])
        ws.cell(row=i, column=2, value=d["orgs"])
        apply_money(ws.cell(row=i, column=3, value=d["fact"]))
        apply_money(ws.cell(row=i, column=4, value=d["plan"]))
        ws.cell(row=i, column=5, value=f"{round(d['fact']/d['plan']*100,1) if d['plan'] else 0}%")
    ws.column_dimensions['A'].width = 30
    for col in [2, 3, 4, 5]: ws.column_dimensions[get_column_letter(col)].width = 18
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=districts_{year}.xlsx"})

@router.get("/export/district")
async def export_district(year: int = Query(default=2022), district: str = Query(...)):
    d = next((x for x in DISTRICTS_DATA if x["name"] == district), {"name": district, "fact": 0, "plan": 0, "orgs": 0})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = district[:30]
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Отчёт по району: {district} ({year} год)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'], ws['B3'] = "Показатель", "Значение"
    style_header(ws, 3, 2)
    data = [("Организаций", d["orgs"]), ("ФАКТ, тыс. ₽", d["fact"]), ("ПЛАН, тыс. ₽", d["plan"]), ("Освоение", f"{round(d['fact']/d['plan']*100,1)}%" if d['plan'] else "0%")]
    for i, (k, v) in enumerate(data, 4):
        ws.cell(row=i, column=1, value=k)
        c = ws.cell(row=i, column=2, value=v)
        if "тыс" in k: apply_money(c)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={district[:20]}_{year}.xlsx"})

@router.get("/history")
async def get_history():
    return [{"filename": "report_2022.xlsx", "date": datetime.now().strftime("%d.%m.%Y"), "status": "success", "records": 274}]

@router.get("/overdue")
async def get_overdue_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Доступ запрещен")

    # Ищем все просроченные отчеты вместе с данными организации
    stmt = (
        select(ReportSubmission)
        .options(joinedload(ReportSubmission.organization))
        .where(ReportSubmission.status == 'overdue')
        .order_by(desc(ReportSubmission.days_overdue)) # Самые злостные должники сверху
    )
    
    result = await db.execute(stmt)
    submissions = result.scalars().all()
    
    return [
        {
            "id": sub.id,
            "organization_name": sub.organization.name,
            "inn": sub.organization.inn,
            "year": sub.year,
            "quarter": sub.quarter,
            "deadline_date": sub.deadline_date,
            "days_overdue": sub.days_overdue
        }
        for sub in submissions
    ]


@router.get("/my-stats")
async def get_my_organization_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "organization" or not current_user.organization_id:
        raise HTTPException(status_code=403, detail="Доступно только организациям")
    
    org_id = current_user.organization_id
    year = date.today().year

    # Ищем факты за этот год
    facts = (await db.execute(select(InvestmentFact).where(
        InvestmentFact.organization_id == org_id, InvestmentFact.year == year
    ))).scalars().all()

    # Ищем план
    forecast = (await db.execute(select(InvestmentForecast).where(
        InvestmentForecast.organization_id == org_id, InvestmentForecast.year == year
    ))).scalars().first()

    # Ищем статус отчетов
    submissions = (await db.execute(select(ReportSubmission).where(
        ReportSubmission.organization_id == org_id, ReportSubmission.year == year
    ))).scalars().all()

    return {
        "year": year,
        "facts": [{"quarter": f.quarter, "amount": f.amount} for f in facts],
        "plan": forecast.forecast_amount if forecast else 0,
        "submissions": [{"quarter": s.quarter, "status": s.status, "deadline": s.deadline_date} for s in submissions]
    }