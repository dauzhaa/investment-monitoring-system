# backend/app/api/routers/monitoring.py
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from app.core.database import get_db
from app.models import Organization, InvestmentReport, District, OKVED
from pydantic import BaseModel
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

router = APIRouter()

class RemindRequest(BaseModel):
    organization_id: int
    year: int
    quarter: int
    email: str

@router.get("/status")
async def get_monitoring_status(
    year: int, 
    quarter: int,  # 0 = весь год, 1-4 = кварталы
    db: AsyncSession = Depends(get_db)
):
    """
    Возвращает статус сдачи отчетности по всем организациям за конкретный период.
    quarter = 0 означает "весь год"
    """
    # 1. Получаем все активные организации
    orgs_res = await db.execute(
        select(Organization)
        .options(
            selectinload(Organization.district),
        )
        .order_by(Organization.name)
    )
    all_orgs = orgs_res.scalars().all()
    
    # 2. Получаем отчеты за выбранный год
    reports_stmt = select(InvestmentReport).where(
        InvestmentReport.year == year
    )
    
    reports_res = await db.execute(reports_stmt)
    reports_by_org = {r.organization_id: r for r in reports_res.scalars().all()}
    
    result = []
    submitted_count = 0
    
    for org in all_orgs:
        report = reports_by_org.get(org.id)
        
        # Определяем, сдан ли отчет в зависимости от квартала
        is_submitted = False
        if report:
            if quarter == 0:  # Весь год
                is_submitted = report.fact_annual > 0
            elif quarter == 1:
                is_submitted = report.fact_q1 > 0
            elif quarter == 2:
                is_submitted = report.fact_q2 > 0
            elif quarter == 3:
                is_submitted = report.fact_q3 > 0
            elif quarter == 4:
                is_submitted = report.fact_q4 > 0
        
        if is_submitted:
            submitted_count += 1
            
        result.append({
            "id": org.id,
            "name": org.name,
            "inn": org.inn,
            "district": org.district.name if org.district else "Не указан",
            "email": org.email,
            "status": "submitted" if is_submitted else "overdue",
            "upload_date": report.created_at.strftime("%Y-%m-%d") if report else None
        })
        
    return {
        "total": len(all_orgs),
        "submitted": submitted_count,
        "percent": round((submitted_count / len(all_orgs)) * 100, 1) if all_orgs else 0,
        "items": result
    }

@router.post("/remind")
async def send_reminder(
    remind_data: RemindRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Отправляет напоминание организации о необходимости сдать отчет.
    В production здесь будет интеграция с email-сервисом.
    """
    org = await db.get(Organization, remind_data.organization_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    
    quarter_text = "весь год" if remind_data.quarter == 0 else f"{remind_data.quarter} квартал"
    
    # TODO: Интеграция с email-сервисом (Celery + SMTP)
    # Здесь должна быть логика отправки email
    print(f"📧 Напоминание отправлено на {remind_data.email}")
    print(f"Организация: {org.name}")
    print(f"Период: {remind_data.year} год, {quarter_text}")
    
    return {
        "status": "success",
        "message": f"Напоминание отправлено на {remind_data.email}"
    }

@router.get("/export/quarterly")
async def export_quarterly_report(
    year: int,
    quarter: int,  # 0 = весь год
    db: AsyncSession = Depends(get_db)
):
    """
    Генерация Excel файла со сводным отчетом за квартал.
    Квартал интерпретируется как накопительный период:
    - 1 кв: январь-март
    - 2 кв: январь-июнь  
    - 3 кв: январь-сентябрь
    - 4 кв: январь-декабрь (весь год)
    """
    # Получаем все организации с их связанными данными
    orgs_res = await db.execute(
        select(Organization)
        .options(
            selectinload(Organization.district),
            selectinload(Organization.okved)
        )
        .order_by(Organization.name)
    )
    all_orgs = orgs_res.scalars().all()
    
    # Получаем отчеты за выбранный год
    reports_res = await db.execute(
        select(InvestmentReport).where(InvestmentReport.year == year)
    )
    reports = {r.organization_id: r for r in reports_res.scalars().all()}
    
    # Создаем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    quarter_text = "Весь год" if quarter == 0 else f"Квартал {quarter}"
    ws.title = f"Отчет {year} {quarter_text}"
    
    # Заголовки
    headers = [
        "Наименование", 
        "ИНН", 
        "Район", 
        "ОКВЭД", 
        "ОКПО", 
        "Почта",
        f"Инвестиции ({quarter_text})",
        "Статус"
    ]
    ws.append(headers)
    
    # Стили заголовков
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные
    for org in all_orgs:
        report = reports.get(org.id)
        
        # Определяем сумму в зависимости от квартала (накопительно)
        investment_amount = 0
        status = "Не сдано"
        
        if report:
            if quarter == 0:  # Весь год
                investment_amount = report.fact_annual
                status = "Сдано" if report.fact_annual > 0 else "Не сдано"
            elif quarter == 1:
                investment_amount = report.fact_q1
                status = "Сдано" if report.fact_q1 > 0 else "Не сдано"
            elif quarter == 2:
                investment_amount = report.fact_q2
                status = "Сдано" if report.fact_q2 > 0 else "Не сдано"
            elif quarter == 3:
                investment_amount = report.fact_q3
                status = "Сдано" if report.fact_q3 > 0 else "Не сдано"
            elif quarter == 4:
                investment_amount = report.fact_q4
                status = "Сдано" if report.fact_q4 > 0 else "Не сдано"
        
        row = [
            org.name,
            org.inn,
            org.district.name if org.district else "",
            org.okved.code if org.okved else "",
            org.okpo or "",
            org.email or "",
            investment_amount,
            status
        ]
        ws.append(row)
        
        # Цветовое выделение статуса
        current_row = ws.max_row
        status_cell = ws.cell(row=current_row, column=8)
        if status == "Не сдано":
            status_cell.font = Font(color="FF0000", bold=True)
        else:
            status_cell.font = Font(color="008000", bold=True)
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    quarter_suffix = "year" if quarter == 0 else f"q{quarter}"
    filename = f"monitoring_{year}_{quarter_suffix}.xlsx"
    
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        output, 
        headers=headers, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.get("/export/organization/{org_id}")
async def export_organization_report(
    org_id: int,
    year: int,
    quarter: int,
    db: AsyncSession = Depends(get_db)
):
    """
    Генерация Excel файла с отчетом конкретной организации.
    """
    org = await db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    
    # Получаем отчет
    report_res = await db.execute(
        select(InvestmentReport).where(
            and_(
                InvestmentReport.organization_id == org_id,
                InvestmentReport.year == year
            )
        )
    )
    report = report_res.scalar_one_or_none()
    
    if not report:
        raise HTTPException(status_code=404, detail="Отчет не найден")
    
    # Создаем Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчет {org.name}"
    
    # Информация об организации
    ws.append(["ОТЧЕТ ОБ ИНВЕСТИЦИЯХ"])
    ws.append([])
    ws.append(["Организация:", org.name])
    ws.append(["ИНН:", org.inn])
    ws.append(["Район:", org.district.name if org.district else ""])
    ws.append(["Год:", year])
    ws.append([])
    
    # Данные по кварталам
    ws.append(["Период", "Сумма инвестиций (млн ₽)"])
    ws.append(["Прогноз (год)", report.forecast_annual])
    ws.append(["1 квартал", report.fact_q1])
    ws.append(["2 квартал", report.fact_q2])
    ws.append(["3 квартал", report.fact_q3])
    ws.append(["4 квартал", report.fact_q4])
    ws.append(["ИТОГО за год", report.fact_annual])
    
    # Стилизация
    title_font = Font(bold=True, size=14)
    ws['A1'].font = title_font
    
    # Автоширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Сохраняем
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"report_{org.inn}_{year}.xlsx"
    
    return StreamingResponse(
        output,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )