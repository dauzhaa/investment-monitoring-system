from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from sqlalchemy.orm import selectinload
from io import BytesIO
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
import logging

from app.core.database import get_db
from app.api.dependencies import get_current_user
from app.core.security import hash_password as get_password_hash

# Модели
from app.models.organization import Organization
from app.models.directories import District
from app.models.investment_fact import InvestmentFact
from app.models.investment_forecast import InvestmentForecast
from app.models.report_submission import ReportSubmission
from app.models.uploaded_file import UploadedFile
from app.models.user import User
from app.models.user_credential import UserCredential
from app.models.audit_log import AuditLog

# Схемы и сервисы
from app.schemas.organization import OrganizationCreate, OrganizationUpdate
from app.services.ipo_calculator import IPOCalculator  # Убедись, что путь правильный

router = APIRouter()
logger = logging.getLogger(__name__)

async def get_organizations_aggregated_data(db: AsyncSession, org_ids: list[int], year: int):
    """
    Эффективная агрегация всех данных для ИПО (Факт, План, Ошибки валидации, Сдачи отчетов).
    """
    if not org_ids:
        return {}, {}, {}, {}

    # 1. Фактические инвестиции
    fact_res = await db.execute(
        select(InvestmentFact.organization_id, func.max(InvestmentFact.amount))
        .where(InvestmentFact.organization_id.in_(org_ids), InvestmentFact.year == year)
        .group_by(InvestmentFact.organization_id)
    )
    facts = dict(fact_res.all())
    
    # 2. Плановые инвестиции (берем последний прогноз по ID)
    plan_subq = select(
        InvestmentForecast.organization_id, 
        func.max(InvestmentForecast.id).label("mid")
    ).where(
        InvestmentForecast.organization_id.in_(org_ids), 
        InvestmentForecast.year == year
    ).group_by(InvestmentForecast.organization_id).subquery()
    
    plan_res = await db.execute(
        select(InvestmentForecast.organization_id, InvestmentForecast.forecast_amount)
        .join(plan_subq, InvestmentForecast.id == plan_subq.c.mid)
    )
    plans = dict(plan_res.all())
    
    # 3. Количество ошибок качества данных (Из UploadedFile)
    errors_res = await db.execute(
        select(UploadedFile.organization_id, func.count(UploadedFile.id))
        .where(
            UploadedFile.organization_id.in_(org_ids),
            UploadedFile.year == year,
            UploadedFile.processing_status == 'error'
        )
        .group_by(UploadedFile.organization_id)
    )
    errors = dict(errors_res.all())

    # 4. История сдач отчетов
    subm_res = await db.execute(
        select(
            ReportSubmission.organization_id, 
            ReportSubmission.quarter, 
            ReportSubmission.status, 
            ReportSubmission.days_overdue
        )
        .where(
            ReportSubmission.organization_id.in_(org_ids), 
            ReportSubmission.year == year
        )
    )
    
    submissions = {org_id: [] for org_id in org_ids}
    for row in subm_res.all():
        submissions[row.organization_id].append({
            "quarter": row.quarter,
            "status": row.status,
            "days_overdue": row.days_overdue
        })
    
    return facts, plans, errors, submissions

def get_current_quarter(target_year: int) -> int:
    """Вычисляет текущий квартал для расчетов."""
    today = date.today()
    if target_year == today.year:
        return (today.month - 1) // 3 + 1
    elif target_year < today.year:
        return 4
    return 1


@router.get("/")
async def get_organizations(
    year: int = Query(default=None), district: Optional[str] = Query(default=None),
    districts: Optional[str] = Query(default=None), smp: Optional[bool] = Query(default=None),
    search: Optional[str] = Query(default=None), 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if year is None: year = date.today().year
    current_quarter = get_current_quarter(year)
        
    stmt = select(Organization).options(selectinload(Organization.district), selectinload(Organization.okved))
    
    if smp is not None: stmt = stmt.where(Organization.is_smp == smp)
    if district: stmt = stmt.join(District).where(District.name == district)
    if districts:
        district_list = [d.strip() for d in districts.split(',') if d.strip()]
        if district_list: stmt = stmt.join(District, isouter=True).where(District.name.in_(district_list))
    if search:
        search_term = f"%{search}%"
        stmt = stmt.where(or_(Organization.name.ilike(search_term), Organization.inn.ilike(search_term)))
    
    stmt = stmt.order_by(Organization.name)
    orgs = (await db.execute(stmt)).scalars().all()
    
    org_ids = [org.id for org in orgs]
    facts, plans, errors, submissions = await get_organizations_aggregated_data(db, org_ids, year)
    
    result = []
    for org in orgs:
        fact_val = float(facts.get(org.id, 0))
        plan_val = float(plans.get(org.id, 0))
        err_count = errors.get(org.id, 0)
        org_subs = submissions.get(org.id, [])
        
        # Получаем статус за 4-й квартал (для обратной совместимости старого поля status)
        q4_status = next((s['status'] for s in org_subs if s['quarter'] == 4), None)

        # Рассчитываем ИПО
        ipo_data = IPOCalculator.calculate(
            is_smp=org.is_smp,
            year=year,
            current_quarter=current_quarter,
            submissions=org_subs,
            errors_count=err_count,
            fact_amount=fact_val,
            plan_amount=plan_val
        )

        result.append({
            "id": org.id,
            "name": org.name,
            "inn": org.inn,
            "district": {"name": org.district.name} if org.district else None,
            "okved": {"code": org.okved.code} if org.okved else None,
            "is_smp": org.is_smp,
            "contact_email": org.contact_email,
            "fact_amount": fact_val,
            "plan_amount": plan_val,
            "status": q4_status,
            "ipo": ipo_data  # Отдаем на фронтенд ИПО и разбивку по индексам
        })
    return result

@router.get("/count")
async def get_organizations_count(db: AsyncSession = Depends(get_db)):
    count = (await db.execute(select(func.count(Organization.id)))).scalar()
    return {"count": count}

@router.get("/export")
async def export_organizations(
    year: int = Query(default=None), districts: Optional[str] = Query(default=None),
    smp: Optional[str] = Query(default=None), db: AsyncSession = Depends(get_db)
):
    if year is None: year = date.today().year
    current_quarter = get_current_quarter(year)
    
    stmt = select(Organization).options(selectinload(Organization.district), selectinload(Organization.okved))
    
    if smp and smp.lower() in ['true', '1']: stmt = stmt.where(Organization.is_smp == True)
    elif smp and smp.lower() in ['false', '0']: stmt = stmt.where(Organization.is_smp == False)
    
    if districts:
        district_list = [d.strip() for d in districts.split(',') if d.strip()]
        if district_list: stmt = stmt.join(District, isouter=True).where(District.name.in_(district_list))
    
    orgs = (await db.execute(stmt.order_by(Organization.name))).scalars().all()
    org_ids = [org.id for org in orgs]
    
    facts, plans, errors, submissions = await get_organizations_aggregated_data(db, org_ids, year)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Организации {year}"
    
    district_names = districts if districts else "Все районы"
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Отчёт об инвестициях и надежности организаций за {year} год"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Районы: {district_names}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Добавили колонки для ИПО
    headers = ['№', 'Наименование', 'ИНН', 'Район', 'ОКВЭД', 'СМП', 'ФАКТ (тыс. ₽)', 'ПЛАН (тыс. ₽)', 'Индекс (ИПО)', 'Статус ИПО']
    header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, org in enumerate(orgs, 5):
        fact_val = float(facts.get(org.id, 0))
        plan_val = float(plans.get(org.id, 0))
        
        # Считаем ИПО для Excel
        ipo_data = IPOCalculator.calculate(
            is_smp=org.is_smp,
            year=year,
            current_quarter=current_quarter,
            submissions=submissions.get(org.id, []),
            errors_count=errors.get(org.id, 0),
            fact_amount=fact_val,
            plan_amount=plan_val
        )

        ws.cell(row=row_num, column=1, value=row_num - 4)
        ws.cell(row=row_num, column=2, value=org.name)
        ws.cell(row=row_num, column=3, value=org.inn)
        ws.cell(row=row_num, column=4, value=org.district.name if org.district else '')
        ws.cell(row=row_num, column=5, value=org.okved.code if org.okved else '')
        ws.cell(row=row_num, column=6, value='Да' if org.is_smp else 'Нет')
        ws.cell(row=row_num, column=7, value=fact_val)
        ws.cell(row=row_num, column=8, value=plan_val)
        ws.cell(row=row_num, column=9, value=ipo_data["ipo"])
        ws.cell(row=row_num, column=10, value=ipo_data["label"])
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"organizations_{year}_filtered.xlsx" if districts else f"organizations_{year}.xlsx"
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/{org_id}")
async def get_organization_details(org_id: int, db: AsyncSession = Depends(get_db)):
    org = await db.get(Organization, org_id)
    if not org: raise HTTPException(status_code=404, detail="Организация не найдена")

    stmt = select(Organization).options(selectinload(Organization.district), selectinload(Organization.okved)).where(Organization.id == org_id)
    org = (await db.execute(stmt)).scalar_one_or_none()

    # Собираем отчеты по годам
    years_res = await db.execute(select(func.distinct(InvestmentFact.year)).where(InvestmentFact.organization_id == org_id))
    years = sorted([y[0] for y in years_res.all()], reverse=True)
    
    reports = []
    for y in years:
        fact_res = await db.execute(select(InvestmentFact.quarter, InvestmentFact.amount).where(InvestmentFact.organization_id == org_id, InvestmentFact.year == y))
        facts_data = {row[0]: float(row[1]) for row in fact_res.all()}
        
        plan_res = await db.execute(select(InvestmentForecast.forecast_amount).where(InvestmentForecast.organization_id == org_id, InvestmentForecast.year == y).order_by(InvestmentForecast.id.desc()).limit(1))
        plan_val = plan_res.scalar_one_or_none() or 0
        
        subm_res = await db.execute(select(ReportSubmission.status).where(ReportSubmission.organization_id == org_id, ReportSubmission.year == y).limit(1))
        status = subm_res.scalar_one_or_none()
        
        reports.append({
            "year": y,
            "fact_annual": facts_data.get(None, max(facts_data.values()) if facts_data else 0),
            "forecast_annual": float(plan_val),
            "fact_q1": facts_data.get(1, 0),
            "fact_q2": facts_data.get(2, 0),
            "fact_q3": facts_data.get(3, 0),
            "fact_q4": facts_data.get(4, 0),
            "status": status
        })

    return {
        "info": {
            "id": org.id, "name": org.name, "inn": org.inn,
            "district": org.district.name if org.district else None,
            "okved": org.okved.code if org.okved else None,
            "is_smp": org.is_smp, "email": org.contact_email
        },
        "reports": reports
    }
    
@router.post("/")
async def create_organization(
    org_in: OrganizationCreate, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Доступ запрещен. Только для администраторов.")

    # Проверка на дубликат ИНН
    existing_org = (await db.execute(select(Organization).where(Organization.inn == org_in.inn))).scalar_one_or_none()
    if existing_org:
        raise HTTPException(status_code=400, detail="Организация с таким ИНН уже существует.")

    try:
        # 1. Создаем организацию
        new_org = Organization(**org_in.model_dump())
        db.add(new_org)
        await db.flush() # Получаем ID организации

        # 2. Создаем пользователя для организации
        org_email = f"info_{new_org.inn}@obr72.ru"
        
        # Проверяем, вдруг юзер с таким email уже есть (осиротевший)
        existing_user = (await db.execute(select(User).where(User.email == org_email))).scalar_one_or_none()
        if not existing_user:
            new_user = User(
                email=org_email,
                role="organization",
                organization_id=new_org.id,
                is_active=True,
                is_email_verified=True
            )
            db.add(new_user)
            await db.flush()
            
            # Создаем пароль
            db.add(UserCredential(
                user_id=new_user.id,
                hashed_password=get_password_hash(new_org.inn) # Пароль = ИНН
            ))

        # 3. Аудит
        audit = AuditLog(
            user_id=current_user.id,
            action="create_organization",
            entity_type="organization",
            entity_id=new_org.id,
            details={"inn": new_org.inn, "name": new_org.name, "created_user_email": org_email}
        )
        db.add(audit)
        
        await db.commit()
        return {"status": "success", "id": new_org.id, "email": org_email, "message": "Организация и доступ созданы."}

    except Exception as e:
        await db.rollback()
        logger.error(f"Error creating organization: {e}")
        raise HTTPException(status_code=500, detail="Ошибка при создании организации.")


@router.put("/{org_id}")
async def update_organization(
    org_id: int, 
    org_in: OrganizationUpdate, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Доступ запрещен.")

    org = await db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")

    update_data = org_in.model_dump(exclude_unset=True) # Только те поля, что переданы
    
    # Запоминаем старые данные для лога
    old_data = {k: getattr(org, k) for k in update_data.keys()}

    for field, value in update_data.items():
        setattr(org, field, value)

    # Аудит
    audit = AuditLog(
        user_id=current_user.id,
        action="update_organization",
        entity_type="organization",
        entity_id=org.id,
        details={"changes": update_data, "old_data": old_data}
    )
    db.add(audit)
    
    await db.commit()
    return {"status": "success", "message": "Организация обновлена"}


@router.delete("/{org_id}")
async def delete_organization(
    org_id: int, 
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Доступ запрещен.")

    org = await db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")

    # Аудит (делаем до удаления, пока есть объект)
    audit = AuditLog(
        user_id=current_user.id,
        action="delete_organization",
        entity_type="organization",
        entity_id=org_id,
        details={"inn": org.inn, "name": org.name}
    )
    db.add(audit)

    await db.delete(org)
    await db.commit()
    
    return {"status": "success", "message": "Организация удалена."}