import asyncio
import openpyxl
from sqlalchemy import select
from app.core.database import AsyncSessionLocal
from app.models.organization import Organization
from app.models.investment_forecast import InvestmentForecast
from app.models.investment_fact import InvestmentFact
from app.models.report_submission import ReportSubmission
from app.models.organization_ipo import OrganizationIPO
from app.services.ipo_calculator import IPOCalculator

async def calculate_and_cache_ipo(year: int = 2024, excel_path: str = '/app/data.xlsx'):
    # 1. ДОБАВЛЕНО: Считываем количество ошибок валидации из нового Excel-файла
    errors_map = {}
    try:
        # data_only=True позволяет считывать результаты формул (если они есть)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        for row_data in ws.iter_rows(min_row=2, values_only=True):
            if not row_data[0]:
                continue
            row = dict(zip(headers, row_data))
            inn = str(row.get('ИНН', '')).strip()
            
            # Безопасно достаем количество ошибок
            errors_raw = row.get('Ошибок валидации')
            try:
                errors = int(float(errors_raw)) if errors_raw is not None else 0
            except (ValueError, TypeError):
                errors = 0
            
            if inn:
                errors_map[inn] = errors
                
        print(f"📊 Успешно загружена статистика ошибок для {len(errors_map)} организаций из Excel.")
    except FileNotFoundError:
        print(f"⚠️ Файл {excel_path} не найден. Все ошибки будут равны 0.")
    except Exception as e:
        print(f"⚠️ Не удалось прочитать данные об ошибках из {excel_path}: {e}")

    # 2. Основной цикл работы с БД
    db = AsyncSessionLocal()
    try:
        orgs = (await db.execute(select(Organization))).scalars().all()
        count = 0
        
        for org in orgs:
            subs = (await db.execute(
                select(ReportSubmission)
                .where(ReportSubmission.organization_id == org.id, ReportSubmission.year == year)
            )).scalars().all()
            submissions_list = [
                {"quarter": s.quarter, "days_overdue": s.days_overdue or 0, "status": s.status} 
                for s in subs
            ]
            
            forecast = (await db.execute(
                select(InvestmentForecast)
                .where(
                    InvestmentForecast.organization_id == org.id, 
                    InvestmentForecast.year == year, 
                    InvestmentForecast.forecast_type == 'первоначальный'
                )
            )).scalars().first()
            plan_amount = float(forecast.forecast_amount) if forecast and forecast.forecast_amount else 0.0
            
            fact = (await db.execute(
                select(InvestmentFact)
                .where(InvestmentFact.organization_id == org.id, InvestmentFact.year == year)
            )).scalars().first()
            fact_amount = float(fact.amount) if fact and fact.amount else 0.0
            
            # 3. ДОБАВЛЕНО: Достаем реальное количество ошибок валидации для этой организации
            actual_errors = errors_map.get(org.inn, 0)
            
            # 4. ИСПРАВЛЕНО: Передаем actual_errors в калькулятор
            ipo_result = IPOCalculator.calculate(
                is_smp=org.is_smp,
                year=year,
                current_quarter=4,
                submissions=submissions_list,
                errors_count=actual_errors,  # <--- Теперь тут не 0!
                fact_amount=fact_amount,
                plan_amount=plan_amount
            )
            
            if ipo_result["ipo"] == "-":
                continue
                
            existing_ipo = (await db.execute(
                select(OrganizationIPO)
                .where(OrganizationIPO.organization_id == org.id, OrganizationIPO.year == year)
            )).scalar_one_or_none()
            
            if existing_ipo:
                existing_ipo.ipo_score = ipo_result["ipo"]
                existing_ipo.d_score = ipo_result["details"]["discipline"]
                existing_ipo.a_score = ipo_result["details"]["quality"]
                existing_ipo.e_score = ipo_result["details"]["execution"]
            else:
                new_ipo = OrganizationIPO(
                    organization_id=org.id,
                    year=year,
                    ipo_score=ipo_result["ipo"],
                    d_score=ipo_result["details"]["discipline"],
                    a_score=ipo_result["details"]["quality"],
                    e_score=ipo_result["details"]["execution"]
                )
                db.add(new_ipo)
            
            count += 1
        
        await db.commit()
        print(f"✅ Успешно рассчитан и сохранен ИПО для {count} организаций за {year} год!")

    except Exception as e:
        await db.rollback()
        print(f"❌ Ошибка БД: {e}")
    finally:
        await db.close()

if __name__ == "__main__":
    # Укажи тут путь к файлу. Если запускаешь локально на Mac, возможно путь будет './data.xlsx'
    # Если в Docker - '/app/data.xlsx'
    asyncio.run(calculate_and_cache_ipo(2024, '/app/data.xlsx'))