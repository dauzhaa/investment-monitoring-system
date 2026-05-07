import openpyxl
import asyncio
from datetime import datetime
from sqlalchemy import select
from app.core.database import AsyncSessionLocal
from app.models.organization import Organization
from app.models.directories import District
from app.models.investment_fact import InvestmentFact
from app.models.investment_forecast import InvestmentForecast
from app.models.report_submission import ReportSubmission


def _to_float(value) -> float:
    """Безопасное преобразование значения из ячейки в float."""
    if value is None or value == '':
        return 0.0
    try:
        return float(str(value).replace(',', '.').replace(' ', ''))
    except (ValueError, TypeError):
        return 0.0


def _clean_email(raw_email: str, inn: str) -> str:
    """Очистка email: берём первый адрес из списка через ; и убираем многоточие."""
    if not raw_email:
        return f"info_{inn}@obr72.ru"
    email = str(raw_email).strip()
    # обрезаем хвост "..." из генерации XLSX
    if '...' in email:
        email = email.split('...')[0]
    # берём первый email из списка через ; , или пробелы
    for sep in [';', ',']:
        if sep in email:
            email = email.split(sep)[0]
            break
    email = email.strip()
    return email if '@' in email else f"info_{inn}@obr72.ru"


def _parse_date(raw):
    """Унифицированный парсинг даты из ячейки Excel."""
    if raw is None or raw == '':
        return None
    if isinstance(raw, datetime):
        return raw.date()
    try:
        return datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
    except ValueError:
        return None


async def import_synthetic_data(xlsx_file_path: str):
    db = AsyncSessionLocal()
    inserted = {'orgs': 0, 'forecasts': 0, 'facts': 0, 'submissions': 0}
    skipped = {'forecasts': 0, 'facts': 0, 'submissions': 0}

    try:
        wb = openpyxl.load_workbook(xlsx_file_path, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for row_data in ws.iter_rows(min_row=2, values_only=True):
            if not row_data[0]:
                continue

            row = dict(zip(headers, row_data))

            # 1. Обработка Района
            district_name = str(row.get('Район', '')).strip()
            if not district_name or district_name == 'None':
                continue

            district = (
                await db.execute(
                    select(District).where(District.name == district_name)
                )
            ).scalar_one_or_none()
            if not district:
                district = District(name=district_name)
                db.add(district)
                await db.flush()

            # 2. Организация
            inn = str(row.get('ИНН', '')).strip()
            if not inn:
                continue

            org = (
                await db.execute(
                    select(Organization).where(Organization.inn == inn)
                )
            ).scalar_one_or_none()

            if not org:
                is_smp_val = (
                    str(row.get('СМП (да/нет)', ''))
                    .strip()
                    .lower() == 'да'
                )
                org = Organization(
                    name=str(row.get('Наименование организации', '')).strip(),
                    inn=inn,
                    district_id=district.id,
                    is_smp=is_smp_val,
                    contact_email=_clean_email(row.get('Email'), inn),
                )
                db.add(org)
                await db.flush()
                inserted['orgs'] += 1

            # 3. Импорт Плана (Forecast) — идемпотентно
            plan_2024 = _to_float(row.get('Прогноз на 2024 год, тыс. руб.'))
            existing_forecast = (
                await db.execute(
                    select(InvestmentForecast).where(
                        InvestmentForecast.organization_id == org.id,
                        InvestmentForecast.year == 2024,
                        InvestmentForecast.forecast_type == 'первоначальный',
                    )
                )
            ).scalar_one_or_none()

            if not existing_forecast:
                db.add(InvestmentForecast(
                    organization_id=org.id,
                    year=2024,
                    forecast_amount=plan_2024,
                    forecast_type='первоначальный',
                ))
                inserted['forecasts'] += 1
            else:
                skipped['forecasts'] += 1

            # 4. Импорт Фактов (InvestmentFact) — идемпотентно
            fact_2024 = _to_float(row.get('Январь-декабрь 2024, тыс. руб.'))
            existing_fact = (
                await db.execute(
                    select(InvestmentFact).where(
                        InvestmentFact.organization_id == org.id,
                        InvestmentFact.year == 2024,
                        InvestmentFact.quarter == 4,
                    )
                )
            ).scalar_one_or_none()

            if not existing_fact:
                db.add(InvestmentFact(
                    organization_id=org.id,
                    year=2024,
                    quarter=4,
                    report_type='годовой',
                    amount=fact_2024,
                ))
                inserted['facts'] += 1
            else:
                skipped['facts'] += 1

            # 5. Мониторинг сдачи (ReportSubmission) — идемпотентно
            existing_sub = (
                await db.execute(
                    select(ReportSubmission).where(
                        ReportSubmission.organization_id == org.id,
                        ReportSubmission.year == 2024,
                        ReportSubmission.quarter == 4,
                    )
                )
            ).scalar_one_or_none()

            if not existing_sub:
                deadline = datetime(2025, 1, 14).date()
                submitted_date = _parse_date(row.get('Дата сдачи отчёта за 2024'))

                if submitted_date is not None:
                    days_overdue = max(0, (submitted_date - deadline).days)
                    status = 'submitted' if days_overdue == 0 else 'overdue'
                    db.add(ReportSubmission(
                        organization_id=org.id,
                        year=2024,
                        quarter=4,
                        deadline_date=deadline,
                        submitted_date=submitted_date,
                        status=status,
                        days_overdue=days_overdue,
                    ))
                    inserted['submissions'] += 1
                else:
                    # организация не сдала отчёт — фиксируем как missing
                    db.add(ReportSubmission(
                        organization_id=org.id,
                        year=2024,
                        quarter=4,
                        deadline_date=deadline,
                        submitted_date=None,
                        status='missing',
                        days_overdue=0,
                    ))
                    inserted['submissions'] += 1
            else:
                skipped['submissions'] += 1

        await db.commit()

        print('Данные из Excel успешно загружены!')
        print(f"  Организации: добавлено {inserted['orgs']}")
        print(
            f"  Прогнозы: добавлено {inserted['forecasts']}, "
            f"пропущено (уже были) {skipped['forecasts']}"
        )
        print(
            f"  Факты: добавлено {inserted['facts']}, "
            f"пропущено (уже были) {skipped['facts']}"
        )
        print(
            f"  Сдачи отчётов: добавлено {inserted['submissions']}, "
            f"пропущено (уже были) {skipped['submissions']}"
        )

    except Exception as e:
        await db.rollback()
        print(f'Ошибка импорта: {e}')
        raise
    finally:
        await db.close()


if __name__ == '__main__':
    asyncio.run(import_synthetic_data('/app/data.xlsx'))